import openpyxl as xl
import pandas as pd
import os

# 获取文件夹内文件路径
def getFile(folderPath):
    fileList = []
    
    for root, dirs, files in os.walk(folderPath):
        for name in files:
            fileList.append(os.path.join(root, name))
    
    return fileList

# 检查文件是否为excel
def checkFile(filePath):
    if '~$' in filePath:
        return False

    exten = ['.xlsx', '.xls']

    return any(filePath.endswith(extension) for extension in exten)

# 读excel数据
def readData(filePath):
    wb = xl.load_workbook(filePath)
    ws = wb['Sheet1']
    rowNum = ws.max_row
    colNum = ws.max_column

    filenName = filePath.split('\\')[-1].split('.')[0]

    datas = [filenName]

    for col in range(1, colNum + 1):
        data = {}
        data['编号'] = col + 1
        data['字段'] = ws.cell(row=1, column=col).value
        data['名称'] = ws.cell(row=3, column=col).value
        data['类型'] = ws.cell(row=2, column=col).value
        data['备注'] = ''

        if ws.cell(row=3, column=col).comment:
            data['备注'] = ws.cell(row=3, column=col).comment.text

        datas.append(data)

    return datas

# 写excel数据
def writeData(datas, writer):
    sheetName = datas[0]
    datas.pop(0)

    pd.DataFrame(datas).to_excel(writer, sheet_name = sheetName, index=False)

if __name__ == "__main__":
    folderPath = os.path.dirname(os.path.realpath(__file__)) + '\\AllData'
    outFile = os.path.dirname(os.path.realpath(__file__)) + '\\outFile.xlsx'

    fileList = getFile(folderPath)

    print('等待...')
    with pd.ExcelWriter(outFile) as writer:
        for filePath in fileList:
            if not checkFile(filePath):
                print('File:%s is not Excel!' % filePath)
                continue

            datas = readData(filePath)

            writeData(datas, writer)

    print('完成')