from xml.etree.ElementTree import Comment
import openpyxl as xl
import os

def getFile(folderPath):
    fileList = []
    
    for root, dirs, files in os.walk(folderPath):
        for name in files:
            fileList.append(os.path.join(root, name))
    
    return fileList

def checkFile(filePath):
    if '~$' in filePath:
        return False

    exten = ['.xlsx', '.xls']

    return any(filePath.endswith(extension) for extension in exten)

def readData(filePath):
    wb = xl.load_workbook(filePath)
    ws = wb['Sheet1']
    rowNum = ws.max_row
    colNum = ws.max_column

    fileName = filePath.split('\\')[-1].split('.')[0]

    datas = [fileName, ['编号','字段','名称','类型','备注']]

    for col in range(1, colNum + 1):
        if ws.cell(row=3,colum=col).comment:
            comment = ws.cell(row=3,colum=col).comment.text
        else:
            comment = ''

        data = [col, ws.cell(row=1,colum=col).value, ws.cell(row=3,colum=col).value, ws.cell(row=2,colum=col).value, comment]

        datas.append(data)

    return datas

def writeData(datas, filePath):
    sheetName = datas[0]
    datas.pop(0)

    wb = xl.load_workbook(filePath)
    ws = wb.create_sheet(sheetName)

    n = 1

    for data in datas:
        ws['A' + str(n)] = data[0]
        ws['B' + str(n)] = data[1]
        ws['C' + str(n)] = data[2]
        ws['D' + str(n)] = data[3]
        ws['E' + str(n)] = data[4]
        n = n + 1

    wb.save(filePath)

if __name__ == "__main__":
    folderPath = os.path.dirname(os.path.realpath(__file__)) + '\\AllData'
    outFile = os.path.dirname(os.path.realpath(__file__)) + '\\outFile.xlsx'

    fileList = getFile(folderPath)

    print('等待...')

    if not os.path.exists(outFile):
        wb = xl.Workbook()
        wb.save(outFile)

    for filePath in fileList:
        if not checkFile(filePath):
            print('File:%s is not Excel!' % filePath)
            continue

        datas = readData(filePath)

        writeData(datas, outFile)

    print('完成')